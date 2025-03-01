
from tqdm.asyncio import tqdm_asyncio
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from tldextract import extract
from difflib import SequenceMatcher
from textacy import preprocessing
from collections import Counter
import pandas as pd
import logging
import warnings
import argparse
import asyncio
import aiohttp
import random
import copy
import ast
import re
import os

def load_proxies(proxy_file):
    """Завантажує список проксі з файлу"""
    proxies = []
    try:
        with open(proxy_file, "r") as f:
            for line in f:
                parts = line.strip().split(":")
                if len(parts) == 4:  # Формат: IP:PORT:USER:PASSWORD
                    ip, port, user, password = parts
                    proxy_url = f"http://{user}:{password}@{ip}:{port}"
                    proxies.append(proxy_url)
    except FileNotFoundError:
        print(f"[WARNING] Файл {proxy_file} не знайдено. Використовуємо запити без проксі.")
    return proxies

def clean_with_nlp(raw_text):
    # Нормалізація пробілів
    text = preprocessing.normalize.whitespace(raw_text)
    # Видалення пунктуації
    text = preprocessing.remove.punctuation(text)
    text = re.sub(r'(\w+)-\s+(\w+)', r'\1\2', text)
    text = re.sub(r'(\S)-\s+(\S)', r'\1\2', text)  # Прибирає переноси типу "revenue-\n recovery"
    text = re.sub(r'\s+', ' ', text)  # Заміна багаторазових пробілів на один
    return text.strip()

def is_similar(sent1, sent2, threshold=0.8):
    return SequenceMatcher(None, sent1, sent2).ratio() > threshold
logging.getLogger('aiohttp').setLevel(logging.ERROR)
warnings.filterwarnings("ignore", category=UserWarning, module='aiohttp')

def find_sentence_with_keyword(text, keyword):
    sentences = re.split(r'(?<=[.!?])\s+', text)
    pattern = r'\b' + re.escape(keyword) + r'\b'
    matching_sentences = [
        sentence.strip() for sentence in sentences
        if re.search(pattern, sentence, flags=re.IGNORECASE)
    ]
    return matching_sentences

def get_previous_sentence(text, target_sentence):
    sentences = re.split(r'(?<=[.!?])\s+', text)
    try:
        index = sentences.index(target_sentence)
        if index > 0:
            return sentences[index - 1].strip()
    except ValueError:
        pass
    return ""

def get_next_sentence(text, target_sentence):
    sentences = re.split(r'(?<=[.!?])\s+', text)
    try:
        index = sentences.index(target_sentence)
        if index < len(sentences) - 1:
            return sentences[index + 1].strip()
    except ValueError:
        pass
    return ""

def str_to_bool(value):
    if isinstance(value, bool):
        return value
    if value.lower() in {'true', '1', 'yes'}:
        return True
    elif value.lower() in {'false', '0', 'no'}:
        return False
    else:
        raise argparse.ArgumentTypeError(f"Invalid boolean value: {value}")

def process_keywords_in_url(df, input_keywords):
    def check_keywords(value):
        matches = [kw for kw in input_keywords if kw in value]
        return bool(matches), matches if matches else None
    df['1 Keyword inside URL'], df['1.1 Keyword in URL'] = zip(*df.iloc[:, 0].astype(str).apply(check_keywords))
    return df


def read_csv_to_pandas(file_path):
    df = pd.read_csv(file_path)
    first_column_name = df.columns[0]
    first_column_df = df[[first_column_name]].copy()

    # ows = first_column_df.iloc[1:500]
    return first_column_df

async def fetch_url(session, url, proxies, timeout=60):
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Connection": "keep-alive",
    }
    proxy = random.choice(proxies) if proxies else None
    try:
        async with session.get(url, headers=headers, timeout=timeout, proxy=proxy) as response:
            if response.status == 200:
                content = await response.text()
                return response.status, content
            else:
                return response.status, None
    except asyncio.TimeoutError:
        return 408, None  # 408: Request Timeout
    except aiohttp.ClientError as e:
        return 502, None  # 502: Bad Gateway або клієнтська помилка
    except Exception as e:
        return 500, None  # 500: Internal Server Error

def filter_by_indices(data, indices_to_remove):
    return [item for index, item in enumerate(data) if index not in indices_to_remove]

def extract_sentences_from_all_tags(url,html_content, keywords, title_keywords, exclude_h_and_true):

    if isinstance(html_content, bytes):
        html_content = html_content.decode('utf-8', errors='ignore')

    soup = BeautifulSoup(html_content, 'html.parser')
    if not soup.body:
        return {
            "kw in text": [],
            "sentence": "",
            "sentence-1": "",
            "sentence+1": "",
            "link_inside_sentence": "",
            "kw in title": [],
            "1.1 kw in title": False,
            "Word count": 0,
            "KC-after": "",
            "KC-before": ""
        }

    # Convert keywords and title_keywords to lowercase
    keywords = [kw.lower() for kw in keywords]
    title_keywords = [kw.lower() for kw in title_keywords]

    for script_tag in soup.find_all("img"):
        script_tag.decompose()

    for script_tag in soup.find_all("noscript"):
        script_tag.decompose()

    for script_tag in soup.find_all("script"):
        script_tag.decompose()
    # Remove all <style> tags with the content
    for style_tag in soup.find_all("style"):
        style_tag.decompose()
    found_keywords = []
    matched_sentences = []
    previous_sentences = []
    next_sentences = []
    link_inside_sentence = []

    # Check for title_keywords in <h1>
    h1_text = " ".join(h1.get_text(separator=" ", strip=True).lower() for h1 in soup.find_all("h1"))
    found_title_keywords = [kw for kw in title_keywords if kw in h1_text]
    has_title_keyword = bool(found_title_keywords)

    # Word count in the filtered body
    body_text = soup.body.get_text(separator=" ", strip=True).lower()
    body_text = clean_with_nlp(body_text)
    word_count = len(body_text.split())

    def get_sibling_text(tag, direction, max_attempts=5):
        """
        Searches for a neighboring text element (next or prev) on the same embedded level.
        If no text is found, continues searching in higher-level elements up to `max_attempts`.
        """
        sibling = tag.find_previous_sibling() if direction == "prev" else tag.find_next_sibling()
        attempts = 0

        while sibling and attempts < max_attempts:
            text = sibling.get_text(strip=True).lower()
            if text:
                # Split text into sentences
                sentences = re.split(r'(?<=[.!?])\s+', text)
                # Return the appropriate sentence
                if direction == "prev":
                    return sentences[-1].strip() if sentences else None
                else:
                    return sentences[0].strip() if sentences else None
            sibling = sibling.find_previous_sibling() if direction == "prev" else sibling.find_next_sibling()
            attempts += 1

        # If nothing is found, check parent siblings
        parent = tag.parent
        while parent and attempts < max_attempts:
            sibling = parent.find_previous_sibling() if direction == "prev" else parent.find_next_sibling()
            while sibling:
                text = sibling.get_text(strip=True).lower()
                if text:
                    sentences = re.split(r'(?<=[.!?])\s+', text)
                    if direction == "prev":
                        return sentences[-1].strip() if sentences else None
                    else:
                        return sentences[0].strip() if sentences else None
                sibling = sibling.find_previous_sibling() if direction == "prev" else sibling.find_next_sibling()
            parent = parent.parent
            attempts += 1

        return None

    for tag_to_remove in soup.find_all(["strong", "em", "b", "i", "u"]):
        tag_to_remove.unwrap()

    for tag in soup.body.find_all(string=True):
        text = tag.strip().lower()
        if not text:
            continue
        for keyword in keywords:
            if keyword in text:
                if tag.parent.name in {"h1", "h2", "h3", "h4", "h5", "h6"}:
                    parent_tag = tag.parent
                    sentence_with_tags = f"<{parent_tag.name}>{parent_tag.get_text(strip=True)}</{parent_tag.name}>"
                    matched_sentences.append(sentence_with_tags)
                    found_keywords.append(keyword)
                    link_inside_sentence.append(tag.parent.name == "a")
                    prev_text = get_sibling_text(parent_tag, "prev")
                    next_text = get_sibling_text(parent_tag, "next")
                    previous_sentences.append(prev_text if prev_text else "")
                    next_sentences.append(next_text if next_text else "")
                    continue

                if tag.parent.name == "a" and tag.parent.parent:
                    parent_tag = tag.parent.parent
                else:
                    parent_tag = tag.parent

                parent_text = parent_tag.get_text(separator=" ", strip=True).lower()
                sentences = re.split(r'(?<=[.!?])\s+', parent_text)

                for i, sentence in enumerate(sentences):
                    if keyword in sentence:
                        for obj in range(sentence.count(keyword)):
                            matched_sentences.append(sentence.strip())
                            found_keywords.append(keyword)
                            link_inside_sentence.append(tag.parent.name == "a")
                            # Ensure the same number of sentences for previous and next
                            prev_sentence = sentences[i - 1].strip() if i > 0 else get_sibling_text(parent_tag, "prev")
                            next_sentence = sentences[i + 1].strip() if i < len(sentences) - 1 else get_sibling_text(parent_tag, "next")
                            previous_sentences.append(prev_sentence if prev_sentence else "")
                            next_sentences.append(next_sentence if next_sentence else "")
    # Extra check if nothin is missing

    for obj in keywords:
        count = len(re.findall(rf'\b{re.escape(obj)}\b', body_text, flags=re.IGNORECASE))
        if count > found_keywords.count(obj):
            count = len(re.findall(rf'\b{re.escape(obj)}\b', body_text, flags=re.IGNORECASE))
            extra_results = find_sentence_with_keyword(body_text, obj)
            for sentence in extra_results:
                if not any(is_similar(sentence[:-2], s) or s in sentence for s in matched_sentences):
                    matched_sentences.append(sentence)
                    found_keywords.append(obj)
                    previous_sentences.append(get_previous_sentence(body_text, sentence))
                    next_sentences.append(get_next_sentence(body_text, sentence))
                    link_inside_sentence.append(False)
    # Ensure equal counts for sentences
    indexes_for_remove = [index for index, value in enumerate(link_inside_sentence) if value]
    indexes_for_remove += [index for index, value in enumerate(matched_sentences) if value.startswith('<h')]

    copy_found_keywords = copy.deepcopy(found_keywords)
    # clear results
    if exclude_h_and_true:
        link_inside_sentence = filter_by_indices(link_inside_sentence, indexes_for_remove)
        found_keywords = filter_by_indices(found_keywords, indexes_for_remove)
        matched_sentences = filter_by_indices(matched_sentences, indexes_for_remove)
        previous_sentences = filter_by_indices(previous_sentences, indexes_for_remove)
        next_sentences = filter_by_indices(next_sentences, indexes_for_remove)

    max_count = max(len(matched_sentences), len(previous_sentences), len(next_sentences))
    previous_sentences.extend([""] * (max_count - len(previous_sentences)))
    next_sentences.extend([""] * (max_count - len(next_sentences)))
    link_inside_sentence.extend([False] * (max_count - len(link_inside_sentence)))

    return {
        "kw in text": found_keywords,
        "sentence": matched_sentences,
        "sentence-1": previous_sentences,
        "sentence+1": next_sentences,
        "link_inside_sentence": link_inside_sentence,
        "kw in title": found_title_keywords,
        "1.1 kw in title": has_title_keyword,
        "Word count": word_count,
        "KC-before": f"{len(copy_found_keywords)}",
        "KC-after": f"{len(found_keywords)}"
    }

async def process_urls_with_keywords(df, input_keywords, title_keywords, exclude_h_and_true,proxies, semaphore_limit=100):
    semaphore = asyncio.Semaphore(semaphore_limit)
    response_status_counts = Counter()

    async def process_url(session, url, keywords, title_keywords):
        async with semaphore:
            status, content = await fetch_url(session, url, proxies)
            response_status_counts[status] += 1
            if content and status == 200:
                result = extract_sentences_from_all_tags(url, content, keywords, title_keywords, exclude_h_and_true)
                result["status"] = status
                return result
            return {"kw in text": "", "sentence": "", "sentence-1": "", "sentence+1": "", "link_inside_sentence": "", "status": status, "kw in title": "", "1.1 kw in title": "","Word count":"","KC-before":"","KC-after":""}

    async with aiohttp.ClientSession() as session:
        tasks = [process_url(session, url, input_keywords, title_keywords) for url in df.iloc[:, 0]]
        results = await tqdm_asyncio.gather(*tasks)
    df["Response Status Code"] = [result["status"] for result in results]
    df["kw in title"] = [result.get("kw in title") for result in results]
    df["1.1 kw in title"] = [result.get("1.1 kw in title") for result in results]
    df["Word count"] = [result.get("Word count") for result in results]
    # df["KC-before"] = [int(result.get("KC-before")) for result in results]
    # df["KC-after"] = [int(result.get("KC-after")) for result in results]
    df["KC-before"] = [int(result.get("KC-before")) if result.get("KC-before") and result.get("KC-before").isdigit() else result.get("KC-before") for result in results]
    df["KC-after"] = [int(result.get("KC-after")) if result.get("KC-after") and result.get("KC-after").isdigit() else result.get("KC-after") for result in results]

    df["Link inside sentence"] = [result["link_inside_sentence"] for result in results]
    df["Keyword in text"] = [result.get("kw in text", []) if result else [] for result in results]
    df["Sentence -1"] = [result["sentence-1"] for result in results]
    df["Sentence"] = [result["sentence"] for result in results]
    df["Sentence +1"] = [result["sentence+1"] for result in results]
    print("\n" + "="*50)
    print("Response Status Code Counts:")
    for status, count in response_status_counts.items():
        print(f"Status {status}: {count} times")
    print("="*50 + "\n")
    return df

def remove_illegal_characters(value):
    try:
        if isinstance(value, list):
            return ", ".join(map(str, value))  # Перетворюємо список на рядок
        if isinstance(value, str):
            return ''.join(char for char in value if ord(char) >= 32 or char == '\n')
        if value is None:
            return ""
        return value
    except IllegalCharacterError as e:
        logging.error(f"{'>'*50}")
        logging.error(e)
        # logging.error(f"{'>'*50}")
        return ""  # Return an empty string if an illegal character is detecte


def clean_cell_value(value):
    if isinstance(value, (int, float)):
        return value  # Directly return numbers without conversion
    if isinstance(value, list):
        value = f"\n{'.'*50}\n".join(map(str, value))  # Convert list to comma-separated string
    elif value is None:
        value = ""
    try:
        value = str(value)  # Convert value to string
        value =  ''.join(char for char in value if ord(char) >= 32 or char == '\n')
    except Exception:
        value = "Invalid Value"
    return value

def get_base_domain(url):
    try:
        extracted = extract(url)
        return f"{extracted.domain}.{extracted.suffix}" if extracted.suffix else extracted.domain
    except Exception:
        return ""

def format_excel_with_dynamic_grouping_and_formulas(df, output_file, fixed_widths=None, delimiter=","):
    """
    Create an Excel file with RAW and Group by domain sheets.
    The Group by domain sheet dynamically calculates values using formulas.
    """
    wb = Workbook()

    # **Sheet 1: RAW**
    ws_raw = wb.active
    ws_raw.title = "RAW"

    # Insert RAW data
    for col_idx, header in enumerate(df.columns, start=1):
        ws_raw.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cleaned_value = clean_cell_value(value)  # Очистка значення
            cell = ws_raw.cell(row=row_idx, column=col_idx, value=cleaned_value)
            
            # Якщо значення число, встановлюємо числовий формат
            if isinstance(cleaned_value, (int, float)):
                cell.number_format = '0'  # Формат для цілих чисел
            elif isinstance(cleaned_value, float):
                cell.number_format = '0.00'  # Формат для чисел з плаваючою точкою


    # **Sheet 2: Group by domain**
    ws_grouped = wb.create_sheet(title="Group by domain")

    df["Domain"] = df["Website"].apply(get_base_domain)

    grouped = df.groupby("Domain")

    # Calculate top keywords dynamically
    keyword_counts = (
        df.explode("Keyword in text")  # Explode lists into rows
        .dropna(subset=["Keyword in text"])  # Remove NaN keywords
        .query("`Keyword in text` != ''")  # Exclude empty strings
        .groupby(["Domain", "Keyword in text"])
        .size()
        .reset_index(name="Count")
    )
    # Get all unique keywords across all domains
    all_keywords = keyword_counts["Keyword in text"].unique().tolist()
    # Dynamic column headers
    # ! do not render [+]
    # headers = ["Domain", "URL Count"] + all_keywords  + list(df.columns)
    # ! end do not render [+]
    headers = ["Domain", "URL Count"] + all_keywords 
    for col_idx, header in enumerate(headers, start=1):
        ws_grouped.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
    # Foud column name
    keyword_col_idx = None
    for col_idx, header in enumerate(df.columns, start=1):
        if header == "Keyword in text":
            keyword_col_idx = col_idx
            break

    if keyword_col_idx is None:
        raise ValueError("Column 'Keyword in text' not found in the DataFrame.")

    # Convert the column index to Excel-style column letters
    keyword_col_letter = get_column_letter(keyword_col_idx)
    # Populate grouped data
    row_idx = 2
    for domain, group in grouped:
        # Add domain row
        ws_grouped.cell(row=row_idx, column=1, value=domain).font = Font(bold=True)

        # **URL Count Formula**
        domain_formula = f'=COUNTIF(RAW!A:A{delimiter}"*{domain}*")'
        ws_grouped.cell(row=row_idx, column=2, value=domain_formula)

        # **Top Keyword Formulas**
        for col_idx, keyword in enumerate(all_keywords, start=3):
            # keyword_formula = f'=COUNTIFS(RAW!A:A{delimiter}"{domain}"{delimiter}RAW!G:G{delimiter}"{keyword}")'
            keyword_formula = (
                f'=SUMPRODUCT((ISNUMBER(SEARCH("{domain}"; RAW!$A:$A))) * '
                f'(LEN(RAW!${keyword_col_letter}:${keyword_col_letter}) - LEN(SUBSTITUTE(RAW!${keyword_col_letter}:${keyword_col_letter}; "{keyword}", ""))) / LEN("{keyword}"))'
            )

            ws_grouped.cell(row=row_idx, column=col_idx, value=keyword_formula)
        # ! do not render [+]
        # # Set group for collapsibility
        # ws_grouped.row_dimensions[row_idx].outlineLevel = 1
        # # Add grouped rows (hidden by default) starting from column 7
        # for _, row in group.iterrows():
        #     row_idx += 1
        #     for col_idx, value in enumerate(row, start=len(all_keywords)+3):
        #         ws_grouped.cell(row=row_idx, column=col_idx, value=clean_cell_value(value))
        #     ws_grouped.row_dimensions[row_idx].outlineLevel = 2
        #     ws_grouped.row_dimensions[row_idx].hidden = True
        # ! end do not render [+]
        # Move to the next domain group
        row_idx += 1

    # Adjust column widths with fixed widths
    for ws in [ws_raw, ws_grouped]:
        for col_idx, col in enumerate(ws.columns, start=1):
            col_letter = get_column_letter(col_idx)
            # if fixed_widths and col_idx - 1 < len(fixed_widths):  # Respect fixed widths
            column_name = ws.cell(row=1, column=col_idx).value
            if column_name in fixed_widths:
                ws.column_dimensions[col_letter].width = fixed_widths[column_name]
            elif fixed_widths:
                # Apply default width (10) for unspecified columns
                ws.column_dimensions[col_letter].width = 20
            else:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col_letter].width = max_length + 2
            # else:
            #     # Auto-adjust if no fixed width specified
            #     max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            #     ws.column_dimensions[col_letter].width = max_length + 2

    # Save the workbook
    wb.save(output_file)
    print(f"File saved to {output_file}")


def main(input_path=None, output_file=None, input_keywords=None, title_keywords=None, exclude_h_and_true=None, proxy_file=None):
    input_path = os.path.abspath(input_path)
    output_file = os.path.abspath(output_file)

    proxies = load_proxies("proxy.txt") if proxy_file else []

    df = read_csv_to_pandas(input_path)
    df = df.rename(columns={df.columns[0]: "Website"})
    df = asyncio.run(process_urls_with_keywords(df, input_keywords, title_keywords, exclude_h_and_true, proxies))
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    fixed_widths = {
        "": 25,
        "Website": 35,
        "Domain":35,
        "Response Status Code": 10,
        "kw in title" : 15,
        "1.1 kw in title": 20,
        "Word count": 10,
        "KC-before": 10,
        "KC-after": 10,
        "Keyword in text": 20,
        "Link inside sentence": 20,
        "Sentence": 70,
        "Sentence -1": 70,
        "Sentence +1": 70
    }
    format_excel_with_dynamic_grouping_and_formulas(df, output_file, fixed_widths=fixed_widths)
    print("*"*50)
    print(f"Results saved to: {output_file}")
    print("*"*50)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Keywords handler")
    parser.add_argument("--input_path", required=True, help="Path to csv file")
    parser.add_argument("--input_keywords", required=True, help="List of keywords if format ['word1', 'word2', ...].")
    parser.add_argument("--title_keywords", required=True, help="List of keywords if format ['word1', 'word2', ...].")
    parser.add_argument("--output_file", required=True, help="Path & Output filename")
    parser.add_argument("--exclude_h_and_true", type=str_to_bool, default=False, help="Exclude entries starting with '<h' or marked as True")
    parser.add_argument("--proxy_file", type=str_to_bool, default=False, help="Path to proxy file")
    args = parser.parse_args()

    try:
        input_keywords = ast.literal_eval(args.input_keywords)
        if not isinstance(input_keywords, list):
            raise ValueError("input_keywords should be a list")
    except Exception as e:
        raise ValueError('Wrong format input_keywords. Use this one ["word1", "word2", ...]') from e

    if args.exclude_h_and_true == True:
        print("."*50)
        print("Excluding turn ON")
        print("."*50)
    elif args.exclude_h_and_true == False:
        print("."*50)
        print("Excluding turn OFF")
        print("."*50)

    main_args = {
        "input_path" : args.input_path,
        "output_file": args.output_file,
        "input_keywords" : ast.literal_eval(args.input_keywords),
        "title_keywords" : ast.literal_eval(args.title_keywords),
        "exclude_h_and_true": args.exclude_h_and_true,
        "proxy_file": args.proxy_file
    }
    main(**main_args)
